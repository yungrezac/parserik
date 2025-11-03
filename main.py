
import re
import requests
import time
import random
import datetime
import math
import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from urllib.parse import urlparse

# --- Основные настройки ---
# Заголовки, маскирующиеся под реальный браузер
headers = {
    'Accept': '*/*',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'keep-alive',
    'Origin': 'https://www.wildberries.ru',
    'Referer': 'https://www.wildberries.ru/',
    'Sec-Ch-Ua-Mobile': '?0',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
}

# --- Вспомогательные функции ---

def make_request(url, headers, timeout=10, retries=5, backoff_factor=0.5):
    """Надежная функция для выполнения HTTP-запросов с повторными попытками."""
    for i in range(retries):
        try:
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()  # Вызовет исключение для кодов 4xx/5xx
            return response
        except requests.exceptions.RequestException as e:
            if isinstance(e, requests.exceptions.HTTPError) and e.response.status_code == 429:
                sleep_time = backoff_factor * (2 ** i) + random.uniform(0, 1)
                print(f"Слишком много запросов. Повторная попытка через {sleep_time:.2f} сек...")
                time.sleep(sleep_time)
                continue
            print(f"Ошибка сети: {e}. Попытка {i + 1} из {retries}...")
            time.sleep(backoff_factor * (2 ** i))
            continue
    raise Exception(f"Не удалось получить данные после {retries} попыток. URL: {url}")

def get_mediabasket_route_map():
    """Получает карту маршрутов для 'корзин' с товарами."""
    try:
        url = 'https://cdn.wbbasket.ru/api/v3/upstreams'
        response = make_request(url, headers, timeout=5)
        data = response.json()
        return data.get('recommend', {}).get('mediabasket_route_map', [{}])[0].get('hosts', [])
    except Exception as e:
        print(f"Предупреждение: Не удалось получить карту маршрутов. {e}")
        return []

def get_host_by_range(value, route_map):
    """Определяет хост 'корзины' по ID товара."""
    if not isinstance(route_map, list): return ''
    for host_info in route_map:
        if host_info.get('vol_range_from', 0) <= value <= host_info.get('vol_range_to', 0):
            return host_info.get('host')
    return ''

def find_value_in_options(options, name):
    """Ищет значение в списке опций товара."""
    if not isinstance(options, list): return ''
    for opt in options:
        if isinstance(opt, dict) and opt.get('name') == name:
            return opt.get('value')
    return ''

def map_data(all_products, columns):
    """Преобразует сырые данные о товарах в структурированный список для Excel."""
    mapped_data = []
    for item in all_products:
        advanced = item.get('advanced', {})
        options = advanced.get('options', [])
        
        row_data = {
            'Артикул продавца': item.get('vendorCode', ''),
            'Бренд': item.get('brand', ''),
            'Наименование': item.get('name', ''),
            'Описание': advanced.get('description', ''),
            'Состав': find_value_in_options(options, 'Состав'),
            'Страна производства': find_value_in_options(options, 'Страна производства'),
            'Комплектация': find_value_in_options(options, 'Комплектация'),
            'ТНВЭД': find_value_in_options(options, 'ТН ВЭД'),
            # --- НОВОЕ ПОЛЕ: КАТЕГОРИЯ ---
            'Категория': find_value_in_options(options, 'Раздел меню') or item.get('subjectName', '')
        }
        
        # Добавляем остальные поля, если они есть в `columns`
        for col_name in columns:
            if col_name not in row_data:
                row_data[col_name] = find_value_in_options(options, col_name)

        mapped_data.append(row_data)
    return mapped_data

def create_excel_file(data, columns, seller_id):
    """Создает Excel-файл из обработанных данных."""
    if not data:
        print("Нет данных для записи в файл.")
        return None

    if not os.path.exists("downloads"):
        os.makedirs("downloads")

    filename = f"seller_{seller_id}_{datetime.datetime.now():%Y-%m-%d_%H-%M}.xlsx"
    output_path = os.path.join("downloads", filename)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Товары продавца {seller_id}"

    # Стили
    header_style = NamedStyle(name="header_style")
    header_style.fill = PatternFill(start_color="9A41FE", end_color="9A41FE", fill_type="solid")
    header_style.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # --- НОВЫЙ ЗАГОЛОВОК ---
    final_columns = ['Артикул продавца', 'Наименование', 'Бренд', 'Категория'] + [col for col in columns if col not in ['Артикул продавца', 'Наименование', 'Бренд', 'Категория']]
    ws.append(final_columns)
    
    for cell in ws[1]:
        cell.style = header_style
    ws.row_dimensions[1].height = 30

    # Данные
    for row_data in data:
        ws.append([row_data.get(header, '') for header in final_columns])

    # Авто-ширина столбцов
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path
    
# --- Основная функция парсинга ---

def run_parser(seller_id, brand_id):
    """Основная логика для сбора и обработки данных о товарах."""
    print("Запуск парсера...")
    all_products = []
    
    print("Получение карты маршрутов WB...")
    baskets = get_mediabasket_route_map()

    # Формируем URL для получения общего количества товаров
    brand_query = f"&fbrand={brand_id}" if brand_id else ""
    url_total_list = f"https://catalog.wb.ru/sellers/v8/filters?ab_testing=false&appType=1&curr=rub&dest=-1257786&supplier={seller_id}{brand_query}&lang=ru&spp=30"

    try:
        res_total = make_request(url_total_list, headers).json()
        products_total = res_total.get('data', {}).get('total', 0)
    except Exception as e:
        print(f"Критическая ошибка при получении общего числа товаров: {e}")
        return

    if not products_total:
        print("Товары не найдены. Проверьте ID продавца и бренда.")
        return

    pages_count = math.ceil(products_total / 100)
    print(f"Найдено товаров: {products_total}. Всего страниц: {pages_count}.")

    # Основной цикл по страницам
    count = 0
    for page_num in range(1, pages_count + 1):
        print(f"Обработка страницы {page_num} из {pages_count}...")
        url_list = f"https://catalog.wb.ru/sellers/v4/catalog?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&page={page_num}&sort=popular&spp=30&supplier={seller_id}{brand_query}"
        
        try:
            products_on_page = make_request(url_list, headers).json().get('products', [])
        except (requests.exceptions.RequestException, json.JSONDecodeError):
            print(f"Не удалось обработать страницу {page_num}. Пропуск.")
            continue

        for item in products_on_page:
            count += 1
            print(f"  [{count}/{products_total}] Обработка товара: {item.get('name', '')[:50]}...")
            
            # Получение детальной информации
            productId = str(item['id'])
            backetName = get_host_by_range(int(productId[:-5]), baskets)
            
            # Если хост не найден, используем перебор
            if not backetName:
                 for i in range(1, 15): # Пробуем стандартные корзины
                    backetFormattedNumber = f"0{i}" if i < 10 else str(i)
                    urlItem = f"https://basket-{backetFormattedNumber}.wbbasket.ru/vol{productId[:-5]}/part{productId[:-3]}/{productId}/info/ru/card.json"
                    try:
                        productResponse = requests.get(urlItem, headers=headers, timeout=2)
                        if productResponse.status_code == 200:
                            item['advanced'] = productResponse.json()
                            break
                    except requests.exceptions.RequestException:
                        continue
                 else: # Если ни одна корзина не подошла
                    item['advanced'] = {}
            else: # Если хост найден
                urlItem = f"https://{backetName}/vol{productId[:-5]}/part{productId[:-3]}/{productId}/info/ru/card.json"
                try:
                    productResponse = requests.get(urlItem, headers=headers, timeout=3)
                    if productResponse.status_code == 200:
                        item['advanced'] = productResponse.json()
                    else:
                        item['advanced'] = {}
                except requests.exceptions.RequestException:
                    item['advanced'] = {}

            all_products.append(item)
            time.sleep(random.uniform(0.1, 0.3)) # Задержка между товарами

        time.sleep(random.uniform(1, 2)) # Задержка между страницами

    print("
Сбор данных завершен. Формирование Excel-файла...")
    
    # Определяем все возможные характеристики для колонок
    all_options = set()
    for p in all_products:
        if 'advanced' in p and 'options' in p['advanced']:
            for opt in p['advanced']['options']:
                all_options.add(opt['name'])
    
    columns = ['Артикул продавца', 'Наименование', 'Бренд', 'Категория', 'Описание', 'Состав', 'Страна производства', 'Комплектация', 'ТНВЭД'] + sorted(list(all_options))

    mapped_data = map_data(all_products, columns)
    
    output_file = create_excel_file(mapped_data, columns, seller_id)
    
    if output_file:
        print(f"\nПарсинг успешно завершен! Файл сохранен в: {output_file}")
    else:
        print("\nНе удалось создать Excel-файл.")

# --- Точка входа ---
if __name__ == '__main__':
    print("--- WB Parser ---")
    seller_id_input = input("Введите ID продавца: ").strip()
    brand_id_input = input("Введите ID бренда (оставьте пустым, если не требуется): ").strip()

    if not seller_id_input.isdigit():
        print("Ошибка: ID продавца должен быть числом.")
    else:
        run_parser(seller_id_input, brand_id_input)
