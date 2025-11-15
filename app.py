
import re
import os
import json
import time
import datetime
import math
import random
from urllib.parse import urlparse, urlencode, parse_qs
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from flask import Flask, render_template, request, send_from_directory, Response, stream_with_context, jsonify

# --- Инициализация Flask ---
app = Flask(__name__, static_folder='public', template_folder='public')
port = int(os.environ.get('PORT', 5000))

# --- Заголовки для запросов ---
headers = {
    'Accept': '*/*',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'keep-alive',
    'Origin': 'https://www.wildberries.ru',
    'Referer': 'https://www.wildberries.ru/',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'cross-site',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
}

# --- Сессия для запросов ---
session = requests.Session()
session.headers.update(headers)

# --- Маршруты API ---
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/categories')
def get_categories():
    try:
        json_path = os.path.join(app.root_path, 'subcategories.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        categories_with_subcategories = {}
        for cat_name, cat_data in data.items():
            if isinstance(cat_data, dict):
                subcategories = {sub_name: sub_data.get('id', '') for sub_name, sub_data in cat_data.items() if isinstance(sub_data, dict)}
                categories_with_subcategories[cat_name] = subcategories
        return jsonify(categories_with_subcategories)

    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "Файл subcategories.json не найден или поврежден."}), 500

@app.route('/stream')
def stream_run():
    seller_id = request.args.get('seller_id')
    brand_id = request.args.get('brand_id')
    xsubject_id = request.args.get('xsubject_id')

    if not seller_id:
        return Response("Ошибка: не указан обязательный параметр (seller_id).", status=400)

    def generate():
        try:
            for progress_update in stream_parser(seller_id, brand_id, xsubject_id):
                yield f"data: {progress_update}\n\n"
                time.sleep(0.05)

        except Exception as e:
            error_payload = json.dumps({"type": "error", "message": f"Критическая ошибка: {e}"}) 
            yield f"data: {error_payload}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/download/<path:filename>')
def download_file(filename):
    directory = os.path.join(os.getcwd(), 'downloads')
    try:
        return send_from_directory(directory, filename, as_attachment=True)
    except FileNotFoundError:
        return "Файл не найден.", 404

# --- Логика парсинга ---
def stream_parser(seller_id, brand_id, xsubject_id=None):
    all_products = []
    seller_name = "Не найден"

    yield json.dumps({'type': 'log', 'message': 'Получение карты маршрутов WB...'})
    baskets = get_mediabasket_route_map()
    if not baskets:
        yield json.dumps({'type': 'log', 'message': 'Не удалось получить карту маршрутов.'})

    brand_query = f"&fbrand={brand_id}" if brand_id else ""
    xsubject_query = f"&xsubject={xsubject_id}" if xsubject_id else ""
    url_total_list = f"https://catalog.wb.ru/sellers/v8/filters?ab_testing=false&appType=1&curr=rub&dest=-1257786&supplier={seller_id}{brand_query}{xsubject_query}&lang=ru&spp=30&uclusters=0"
    
    try:
        res_total = make_request(url_total_list).json()
        products_total = res_total.get('data', {}).get('total', 0)
    except Exception as e:
        raise Exception(f"Ошибка при получении общего числа товаров: {e}")

    if not products_total:
        raise Exception("Товары не найдены. Проверьте ID продавца, бренда или подкатегории.")

    pages_count = math.ceil(products_total / 100)
    yield json.dumps({'type': 'start', 'total': products_total, 'seller_name': seller_name, 'message': f'Найдено товаров: {products_total}.'})

    count = 0
    for page_num in range(1, pages_count + 1):
        url_list = f"https://catalog.wb.ru/sellers/v4/catalog?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&page={page_num}&sort=popular&spp=30&supplier={seller_id}{brand_query}{xsubject_query}"
        try:
            products_on_page = make_request(url_list).json().get('products', [])
        except (requests.exceptions.RequestException, json.JSONDecodeError):
            continue

        for item in products_on_page:
            count += 1
            yield json.dumps({'type': 'progress', 'current': count, 'total': products_total, 'message': item.get('name', '')})
            
            productId = str(item['id'])
            backetName = get_host_by_range(int(productId[:-5]), baskets)
            backetNumber, isAutoServer = 1, bool(backetName)
            while True:
                if not isAutoServer and backetNumber > 12: item['advanced'] = {}; break
                backetFormattedNumber = f"0{backetNumber}" if backetNumber < 10 else str(backetNumber)
                urlItem = f"https://{backetName if isAutoServer else f'basket-{backetFormattedNumber}.wbbasket.ru'}/vol{productId[:-5]}/part{productId[:-3]}/{productId}/info/ru/card.json"
                try:
                    productResponse = make_request(urlItem, timeout=3)
                    if productResponse.status_code == 200: item['advanced'] = productResponse.json(); break
                    if not isAutoServer and productResponse.status_code == 404: backetNumber += 1; continue
                    item['advanced'] = {}; break
                except requests.exceptions.RequestException: item['advanced'] = {}; break
            
            all_products.append(item)
            time.sleep(random.uniform(0.1, 0.3))

        time.sleep(random.uniform(1, 2))

    columns = get_columns_for_subcategory(xsubject_id)

    yield json.dumps({'type': 'log', 'message': 'Формирование таблицы...'})
    mapped_data = map_data(all_products, columns)

    yield json.dumps({'type': 'log', 'message': 'Создание Excel-файла...'})
    output_path = create_excel_file(mapped_data, columns)
    if not output_path:
        raise Exception("Не удалось создать Excel-файл.")

    download_filename = os.path.basename(output_path)
    yield json.dumps({'type': 'result', 'download_filename': download_filename, 'total': products_total})

def get_columns_for_subcategory(xsubject_id):
    if not xsubject_id:
        return []
    json_path = os.path.join(app.root_path, 'subcategories.json')
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for cat_data in data.values():
        if isinstance(cat_data, dict):
            for sub_data in cat_data.values():
                if isinstance(sub_data, dict) and sub_data.get('id') == xsubject_id:
                    return sub_data.get('columns', [])
    return []

def make_request(url, timeout=20, retries=10, backoff_factor=1):
    for i in range(retries):
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException:
            time.sleep(backoff_factor * (2 ** i) + random.uniform(0, 1))
            continue
    raise Exception(f"Не удалось получить данные после {retries} попыток. URL: {url}")

def get_mediabasket_route_map():
    try:
        r = make_request('https://cdn.wbbasket.ru/api/v3/upstreams', timeout=5)
        return r.json().get('recommend', {}).get('mediabasket_route_map', [{}])[0].get('hosts', [])
    except Exception: return []

def get_host_by_range(val, route_map):
    if not isinstance(route_map, list): return ''
    for host in route_map: 
        if host.get('vol_range_from') <= val <= host.get('vol_range_to'): return host.get('host')
    return ''

def map_data(data, columns):
    master_mapping = {
        'Артикул продавца': lambda item, adv: item.get('vendorCode', ''),
        'Бренд': lambda item, adv: item.get('brand', ''),
        'Наименование': lambda item, adv: item.get('name', ''),
        'Описание': lambda item, adv: adv.get('description', ''),
        'Состав': lambda item, adv: find_value_in_options(adv.get('options', []), 'Состав'),
        'Страна производства': lambda item, adv: find_value_in_options(adv.get('options', []), 'Страна производства'),
        'Комплектация': lambda item, adv: find_value_in_options(adv.get('options', []), 'Комплектация'),
        'ТНВЭД': lambda item, adv: find_value_in_options(adv.get('options', []), 'ТН ВЭД'),
        'Категория продавца': lambda item, adv: adv.get('subj_root_name', ''),
    }
    new_data = []
    for item in data:
        advanced = item.get('advanced', {})
        row_data = {}
        for col_name in columns:
            if col_name in master_mapping:
                row_data[col_name] = master_mapping[col_name](item, advanced)
            else:
                row_data[col_name] = find_value_in_options(advanced.get('options', []), col_name)
        new_data.append(row_data)
    return new_data

def find_value_in_options(options, name):
    if not isinstance(options, list): return ''
    for opt in options: 
        if isinstance(opt, dict) and opt.get('name') == name: return opt.get('value')
    return ''

def create_excel_file(data, columns):
    if not data: return None
    if not os.path.exists("downloads"): os.makedirs("downloads")
    
    filename = f"wb_parse_{datetime.datetime.now():%Y-%m-%d_%H-%M}.xlsx"
    output_path = os.path.join("downloads", filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    header_style = NamedStyle(name="header_style")
    header_style.fill = PatternFill(start_color="6A5ACD", end_color="6A5ACD", fill_type="solid")
    header_style.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(header_style)
    
    if columns:
        ws.append(columns)
        for cell in ws[1]:
            cell.style = header_style
    
    for row_data in data:
        ws.append([row_data.get(header, '') for header in columns])
        
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3) if max_length < 50 else 50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=port, debug=True)
